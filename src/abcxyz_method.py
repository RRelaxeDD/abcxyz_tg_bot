import matplotlib.pyplot as pyplot
import matplotlib
import openpyxl
import io
from string import ascii_uppercase
from datetime import datetime



def abcmethod(file) -> [io.BytesIO, io.BytesIO]:

        """
        make abcxyz method
        """

        # open excel file
        wb = openpyxl.open(io.BytesIO(file))
        wb.active = 0
        sheet = wb.active

        all_rows = tuple(wb.active.rows)

        # rows count
        lenoflist = len(all_rows)
        
        # columns count, max 16
        column_is = len(all_rows[0]) -1
        # print(column_is)
        

        # variables to store new data
        perem_end_sum_is = ascii_uppercase[column_is+1]
        perem_part_from_sum = ascii_uppercase[column_is+2]
        perem_sovocup_percent = ascii_uppercase[column_is+3]
        perem_abc = ascii_uppercase[column_is+4]
        perem_mid_value = ascii_uppercase[column_is+5]
        perem_sq_value = ascii_uppercase[column_is+6]
        perem_koef_var = ascii_uppercase[column_is+7]
        perem_xyz = ascii_uppercase[column_is+8]
        perem_abcxyz = ascii_uppercase[column_is+9]
        
        # Форматирование файла
        sheet.merge_cells(f'{perem_end_sum_is}1:{perem_end_sum_is}2')
        sheet[f'{perem_end_sum_is}1'] = 'Итог'
        
        sheet.merge_cells(f'{perem_part_from_sum}1:{perem_part_from_sum}2')
        sheet[f'{perem_part_from_sum}1'] = 'Доля в обороте'

        sheet.merge_cells(f'{perem_sovocup_percent}1:{perem_sovocup_percent}2')
        sheet[f'{perem_sovocup_percent}1'] = 'Совокупный процент'

        sheet.merge_cells(f'{perem_abc}1:{perem_abc}2')
        sheet[f'{perem_abc}1'] = 'ABC'

        sheet.merge_cells(f'{perem_mid_value}1:{perem_mid_value}2')
        sheet[f'{perem_mid_value}1'] = 'Средее значение'

        sheet.merge_cells(f'{perem_sq_value}1:{perem_sq_value}2')
        sheet[f'{perem_sq_value}1'] = 'Среднеквадратичное отклонение'

        sheet.merge_cells(f'{perem_koef_var}1:{perem_koef_var}2')
        sheet[f'{perem_koef_var}1'] = 'Коэффицент вариации'

        sheet.merge_cells(f'{perem_xyz}1:{perem_xyz}2')
        sheet[f'{perem_xyz}1'] = 'XYZ'

        sheet.merge_cells(f'{perem_abcxyz}1:{perem_abcxyz}2')
        sheet[f'{perem_abcxyz}1'] = 'ABCXYZ'


        # calculate sum of all values and for each
        sum_of_all = 0
        list_of_values = []

        for num, i in enumerate(all_rows[2:]):
            sum_for_one = sum((i.value for i in i[1:]))

            sum_of_all += sum_for_one
            sheet[f'{perem_end_sum_is}{num + 3}'] = sum_for_one
            list_of_values.append(sum_for_one)



        # calculate percentage from sum_of_all for each
        tmpindex = 3
        part_from_all_for_one_forsovocup = []

        for i in list_of_values:
            partva = i / sum_of_all
            partval = partva * 100
            sheet[f'{perem_part_from_sum}{tmpindex}'] = partval
            tmpindex += 1


        ####
        # sorting
        iforsort1 = 3
        list_to_sort = []

        for i in range(3, lenoflist+1):

            list_values_forone = []
            list_values_forone.append(sheet[f'{perem_part_from_sum}{iforsort1}'].value)
            list_of_values.append(sheet[f'{perem_end_sum_is}{iforsort1}'].value)

            for j in ascii_uppercase[:column_is]:
                list_values_forone.append(sheet[f'{j}{i}'].value)

            list_values_forone.append(sheet[f'{perem_end_sum_is}{i}'].value)
            list_to_sort.append(list_values_forone)
            iforsort1 += 1

        list_to_sort.sort()
        list_to_sort = list_to_sort[::-1] # ! dont change

        tmpindex = 3
        for i in list_to_sort:

            sheet['A'+str(tmpindex)] = i[1]
            sheet[f'{perem_end_sum_is}{tmpindex}'] = i[-1]
            sheet[f'{perem_part_from_sum}{tmpindex}'] = i[0]
            summonths = i[2:-2]

            for j, k in zip(ascii_uppercase[1:column_is], summonths):
                
                sheet[f'{j}{tmpindex}'] = k

            part_from_all_for_one_forsovocup.append(i[0])
            tmpindex += 1
        ####


        # calculating sovocup percent
        sovocup_percent = []
        sovocup_percent.append(list_to_sort[0][0])
        j = 1
        for i in range(0, len(part_from_all_for_one_forsovocup)):
            if j >= len(part_from_all_for_one_forsovocup):
                j = 0
                break
            plisprev = sovocup_percent[i]+list_to_sort[j][0]
            sovocup_percent.append(plisprev)
            j += 1


        for num, i in enumerate(sovocup_percent):
            sheet[f'{perem_sovocup_percent}{num+3}'] = i


        # define product category by sovocup percent
        category = []
        for i in range(0, len(sovocup_percent)):
            if sovocup_percent[i] <= 80:
                category.append('A')
            elif 80 < sovocup_percent[i] <= 90:
                category.append('B')
            else:
                category.append('C')


        for num, i in enumerate(category):
            sheet[f'{perem_abc}{num+3}'] = i


        # calculate xyz
        sq_value = []
        mid_money = []
        koef_value = []
        xyz_value = []
        abc_xyz = []


        # calculate mid and min sq value
        for i in range(3, lenoflist+1):

            plusthing = 0
            list_of_values_forsq = []

            for j in ascii_uppercase[1:column_is]:

                plusthing += sheet[f'{j}{i}'].value
                list_of_values_forsq.append(sheet[f'{j}{i}'].value)

            dividething = plusthing/column_is
            mid_money.append(dividething)
            list_of_values_forsq2 = []

            for i in list_of_values_forsq:

                list_of_values_forsq2.append((i-dividething)**2)

            plussecondthing = sum(list_of_values_forsq2)
            seconddividething = plussecondthing/column_is
            mid_square = seconddividething**(1/2)

            sq_value.append(mid_square)


        # display mid value
        for num, i in enumerate(mid_money):
            sheet[f'{perem_mid_value}{num+3}'] = i


        # display mid sq value
        for num, i in enumerate(sq_value):
            sheet[f'{perem_sq_value}{num+3}'] = i


        # calculate and display koef
        for i in range(3, lenoflist+1):
            res_koef = sq_value[i-3]/mid_money[i-3]*100
            koef_value.append(res_koef)
            sheet[f'{perem_koef_var}{i}'] = res_koef


        # define category and display it
        for num, i in enumerate(koef_value):

            if i < 10:
                res_xyz = 'X'
            
            elif 25 < i:
                res_xyz = 'Z'
            
            else:
                res_xyz = 'Y'

            xyz_value.append(res_xyz)
            sheet[f'{perem_xyz}{num + 3}'] = res_xyz


        # concatinate abc and xyz
        for i in range(3, lenoflist+1):
            xyz_abc = category[i-3]+xyz_value[i-3]
            abc_xyz.append(xyz_abc)
            sheet[f'{perem_abcxyz}{i}'] = xyz_abc


        # save new file with result near original file
        new_file = io.BytesIO()
        # new_file.
        "yyyy.mm.dd-12-12"
        # wb.save
        wb.save(new_file)
        wb.close()
        new_file = io.BytesIO(new_file.getvalue())
        new_file.name = f"abcxyz_method_{str(datetime.now())[:16].replace('-', '_').replace(' ', '_').replace(':', '_')}.xlsx"



        # making diagram


        alpha_forallthings = category

        display_range_a = []
        display_range_adigit = []

        display_range_b = []
        display_range_bdigit = []

        display_range_c = []
        display_range_cdigit = []

        display_x_a_digit = []
        display_x_b_digit = []
        display_x_c_digit = []

        # Добавляет данные в листы для отображения диаграммы
        for i, j, k in zip(alpha_forallthings, list_to_sort, koef_value):
            if i == 'A':
                display_range_a.append(i)
                display_range_adigit.append(j[0])
                display_x_a_digit.append(k)

            if i == 'B':
                display_range_b.append(i)
                display_range_bdigit.append(j[0])
                display_x_b_digit.append(k)

            if i == 'C':
                display_range_c.append(i)
                display_range_cdigit.append(j[0])
                display_x_c_digit.append(k)

        # Создает отметки на диаграмме
        y_things1 = ['', 'A', '', 'B', '', 'C']
        x_things1 = ['X', '', 'Y', '', 'Z', '']
        y_things2 = [
                        display_range_adigit[0], display_range_adigit[len(display_range_adigit)//2],  
                        display_range_bdigit[0], display_range_bdigit[len(display_range_bdigit)//2],
                        display_range_cdigit[0], display_range_cdigit[len(display_range_cdigit)//2]
                    ]
        x_things2 = [5, 10, 17, 25, (max(koef_value)-25)/2+25, max(koef_value)]

        # Добавляет данные на диаграмму
        # pyplot.
        matplotlib.use('agg')

        pyplot.grid()
        pltt = pyplot
        pltt.scatter(x=display_x_a_digit, y=display_range_adigit, s=4, c='red')
        pltt.scatter(display_x_b_digit, display_range_bdigit, s=4, c='green')
        pltt.scatter(display_x_c_digit, display_range_cdigit, s=4, c='blue')
        pltt.yticks(y_things2, y_things1)
        pltt.xticks(x_things2, x_things1)

        diagram = io.BytesIO()
        pltt.savefig(diagram, format='pdf')
        diagram = io.BytesIO(diagram.getvalue())
        diagram.name =  f"abcxyz_method_{str(datetime.now())[:16].replace('-', '_').replace(' ', '_').replace(':', '_')}.pdf"
        pltt.close()
        return [new_file, diagram]
